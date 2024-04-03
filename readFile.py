
#print(file.read()) # read all the contents of file
# print(file.read(5)) # read n number of characters by passing parameter(xuống dòng tính 1 kí tự)
#print("---------------")
#print(file.readline()) # read one single line at a time readLine()
# print line by line using readline method
#print(file.readline()) # mỗi câu lệnh đọc 1 dòng, gọi lần 2 là đọc dòng 2

# print line by line using readline method
#line = file.readline()
#while line!="":   #not blank
#    print(line)
#    line = file.readline()

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

#--Chrome
service_obj = Service("C:/Users/HP/OneDrive - PowerGate/Documents/chromedriver.exe") # seleniumManager
options = Options()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=options, service=service_obj)
driver.get("https://tinhte.vn/forums/xe-hoi.545/")
title1 = driver.find_element(By.XPATH,"/html/body/div[2]/div[1]/div[2]/div[2]/div[3]/div/div/div/div/div/div[4]/form[1]/ol/li[1]/div[2]/div/h3/a").text
driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[3]/div/div/div/div/div/div[4]/form[1]/ol/li[1]/div[2]/div/h3/a").click()
content1 = driver.find_element(By.CSS_SELECTOR,"span.xf-body-paragraph").text
#values = [abc, bdbgfh,"cat", dog, elephant]
# with open('Paths.txt','r') as reader:
#     for locator in reader.readlines():
#         print(locator)

# with open('content.txt','w') as writer:
#         for line in reader.readlines():
#             writer.write(line)
from openpyxl import Workbook
wb = Workbook()
sheet = wb.active
sheet.title = "crawl data"
c1 = sheet.cell(row = 1, column = 1)
c1.value = title1
c2 = sheet.cell(row= 1 , column = 2)
c2.value = content1
# Once have a Worksheet object, one can
# access a cell object by its name also.
# A2 means column = 1 & row = 2.
c3 = sheet['A2']
c3.value = "RAHUL"

# B2 means column = 2 & row = 2.
c4 = sheet['B2']
c4.value = "RAI"

# Anytime you modify the Workbook object
# or its sheets and cells, the spreadsheet
# file will not be saved until you call
# the save() workbook method. 
wb.save("C:\\Users\\HP\\PycharmProjects\\content.xlsx")


